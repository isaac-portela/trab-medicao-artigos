from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.ifrd import classify_ifrd
from src.models import ArticleResult, ProcessingStatus, SynthesisOutput


MAIN_COLUMNS = [
    "Grupo",
    "Aluno",
    "Título",
    "Pasta",
    "Link",
    "ProcessingStatus",
    "Erro",
    "IFRD",
    "Classificação IFRD",
    "Citações",
    "Origem Citações",
    "Tema",
    "Ano Publicação",
    "Editora/Entidade",
    "Tipo Veículo",
    "Tipo Estudo",
    "Natureza Pesquisa",
    "Natureza Dados",
    "Tamanho Amostra",
    "Métodos Estatísticos",
    "Métricas Software",
    "Unidades Plano",
    "Tópicos Específicos",
    "Resumo PT",
    "Questão de Pesquisa",
    "Metodologia",
    "Achados do Reader",
    "Técnicas Estatísticas Reader",
    "Nota: Qualidade",
    "Nota: Replicabilidade",
    "Nota: Aplicabilidade",
    "Nota: Contribuição Teórica",
    "Nota: Adequação",
    "Nota: Aprendizagem",
    "Nota: Alinhamento",
    "Justificativa: Qualidade",
    "Justificativa: Replicabilidade",
    "Justificativa: Aplicabilidade",
    "Justificativa: Contribuição Teórica",
    "Justificativa: Adequação",
    "Justificativa: Aprendizagem",
    "Justificativa: Alinhamento",
    "Disponibilidade Dados/Código",
    "Links Replicação",
    "Artefatos Compartilhados",
    "3 Principais Descobertas",
    "Principal Limitação",
]


class ExcelExporter:
    @staticmethod
    def export(results: list[ArticleResult], synthesis: SynthesisOutput | None, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        rows = [ExcelExporter._row_for_result(result) for result in results]
        df = pd.DataFrame(rows, columns=MAIN_COLUMNS)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Artigos Classificados", index=False)
            workbook = writer.book
            main_sheet = writer.sheets["Artigos Classificados"]
            ExcelExporter._style_main_sheet(main_sheet, df)
            ExcelExporter._write_synthesis_sheet(workbook, synthesis)

    @staticmethod
    def _join(values: list[Any] | None) -> str:
        if not values:
            return ""
        return "\n".join(str(value) for value in values)

    @staticmethod
    def _row_for_result(result: ArticleResult) -> dict[str, Any]:
        metadata = result.metadata
        classification = result.classification
        critic = result.critic_output
        summary = result.super_summary
        ifrd_label = classify_ifrd(result.ifrd)[0] if result.ifrd is not None else ""

        citation_count = result.citation_count
        citation_source = result.citation_source

        # Resolve citation source/count dynamically if missing or legacy
        if citation_source is None or citation_source in ("N/A", "cache"):
            from src.services.citation_lookup import CitationLookupService
            try:
                service = CitationLookupService()
                citation_res = service.get_citation_count(metadata.title)
                citation_source = citation_res.source
                if citation_res.count is not None:
                    citation_count = citation_res.count
            except Exception:
                pass

        row = {
            "Grupo": metadata.group,
            "Aluno": metadata.student,
            "Título": metadata.title,
            "Pasta": str(metadata.folder_path),
            "Link": metadata.link,
            "ProcessingStatus": result.processing_status.value,
            "Erro": result.error or "",
            "IFRD": result.ifrd,
            "Classificação IFRD": ifrd_label,
            "Citações": citation_count if citation_count is not None else "N/A",
            "Origem Citações": citation_source if citation_source is not None else "N/A",
            "Tema": classification.theme if classification else "",
            "Ano Publicação": classification.publication_year if classification else "",
            "Editora/Entidade": classification.publisher_entity if classification else "",
            "Tipo Veículo": classification.venue_type if classification else "",
            "Tipo Estudo": classification.study_type if classification else "",
            "Natureza Pesquisa": classification.research_nature if classification else "",
            "Natureza Dados": classification.data_nature if classification else "",
            "Tamanho Amostra": classification.sample_size if classification else "",
            "Métodos Estatísticos": ExcelExporter._join(classification.statistical_methods if classification else []),
            "Métricas Software": ExcelExporter._join(classification.software_metrics if classification else []),
            "Unidades Plano": ExcelExporter._join(classification.syllabus_units if classification else []),
            "Tópicos Específicos": ExcelExporter._join(classification.syllabus_topics if classification else []),
            "Resumo PT": critic.resumo_pt if critic else "",
            "Questão de Pesquisa": summary.core_research_question if summary else "",
            "Metodologia": summary.methodology_description if summary else "",
            "Achados do Reader": ExcelExporter._join(summary.key_findings if summary else []),
            "Técnicas Estatísticas Reader": ExcelExporter._join(summary.statistical_techniques if summary else []),
            "Disponibilidade Dados/Código": critic.data_availability if critic else "",
            "Links Replicação": ExcelExporter._join(critic.replication_links if critic else []),
            "Artefatos Compartilhados": ExcelExporter._join(critic.shared_artifacts if critic else []),
            "3 Principais Descobertas": ExcelExporter._join(critic.main_findings if critic else []),
            "Principal Limitação": critic.principal_limitation if critic else "",
        }

        rubric_fields = {
            "qualidade_academica": "Qualidade",
            "replicabilidade": "Replicabilidade",
            "aplicabilidade_pratica": "Aplicabilidade",
            "contribuicao_teorica": "Contribuição Teórica",
            "adequacao_aluno": "Adequação",
            "contribuicao_aprendizagem": "Aprendizagem",
            "alinhamento_plano": "Alinhamento",
        }
        for field_name, label in rubric_fields.items():
            score = getattr(critic, field_name, None) if critic else None
            row[f"Nota: {label}"] = score.score if score else ""
            row[f"Justificativa: {label}"] = score.justification if score else ""

        return row

    @staticmethod
    def _style_main_sheet(worksheet, df: pd.DataFrame) -> None:
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        fills = {
            "green": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
            "yellow": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "red": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
            "orange": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        }

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        status_col = df.columns.get_loc("ProcessingStatus") + 1
        ifrd_col = df.columns.get_loc("IFRD") + 1
        classification_col = df.columns.get_loc("Classificação IFRD") + 1

        for row_idx in range(2, worksheet.max_row + 1):
            status = worksheet.cell(row=row_idx, column=status_col).value
            ifrd_value = worksheet.cell(row=row_idx, column=ifrd_col).value
            row_fill = None
            if status and status != ProcessingStatus.SUCCESS.value:
                row_fill = fills["orange"]
            elif isinstance(ifrd_value, (int, float)):
                row_fill = fills[classify_ifrd(float(ifrd_value))[1]]

            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_fill:
                    cell.fill = row_fill
            if isinstance(ifrd_value, (int, float)):
                fill = fills[classify_ifrd(float(ifrd_value))[1]]
                worksheet.cell(row=row_idx, column=ifrd_col).fill = fill
                worksheet.cell(row=row_idx, column=classification_col).fill = fill

        worksheet.freeze_panes = "A2"
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "")
            column_letter = get_column_letter(column_cells[0].column)
            if header in {"Resumo PT", "Metodologia", "Achados do Reader", "3 Principais Descobertas", "Erro"}:
                width = 45
            elif header in {"Título", "Pasta", "Link", "Tópicos Específicos"}:
                width = 38
            else:
                width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 10), 30)
            worksheet.column_dimensions[column_letter].width = width

    @staticmethod
    def _write_synthesis_sheet(workbook, synthesis: SynthesisOutput | None) -> None:
        sheet = workbook.create_sheet("Síntese Temática")
        sheet.append(["Seção", "Tipo/Label", "Referências", "Descrição"])
        if synthesis is None:
            sheet.append(["Síntese", "", "", "Síntese não gerada ou indisponível."])
        else:
            for relationship in synthesis.relationships:
                sheet.append(
                    [
                        "Relação temática",
                        relationship.relationship_type,
                        ExcelExporter._join(relationship.article_references),
                        relationship.description,
                    ]
                )
            for cluster in synthesis.clusters:
                sheet.append(["Cluster", cluster.label, ExcelExporter._join(cluster.article_references), ""])
            discussion = synthesis.ifrd_discussion
            sheet.append(["IFRD", "Racional", "", discussion.formula_rationale])
            sheet.append(["IFRD", "Limitações", "", ExcelExporter._join(discussion.limitations)])
            sheet.append(["IFRD", "Conclusão", "", discussion.viability_conclusion])
            statements = [
                ("Sem overlap", synthesis.no_overlap_statement),
                ("Sem continuidade", synthesis.no_continuity_statement),
                ("Sem complementação", synthesis.no_complementation_statement),
            ]
            for label, statement in statements:
                if statement:
                    sheet.append(["Declaração", label, "", statement])

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"
        for idx, width in enumerate([22, 24, 42, 90], start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

