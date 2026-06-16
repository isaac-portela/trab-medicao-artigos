import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, Any

class ExcelExporter:
    """Class responsible for calculations and exporting the classification data to a formatted Excel file."""

    @staticmethod
    def calculate_ifrd(row: Dict[str, Any]) -> float:
        """Calculates the IFRD index based on the rubric scores."""
        # IFRD = 0.25*Qualidade + 0.20*Alinhamento + 0.20*Aprendizagem + 0.15*Replicabilidade + 0.10*Aplicabilidade + 0.10*Adequacao
        q = row.get("Nota: Qualidade", 3.0)
        ali = row.get("Nota: Alinhamento", 3.0)
        apr = row.get("Nota: Aprendizagem", 3.0)
        rep = row.get("Nota: Replicabilidade", 3.0)
        app = row.get("Nota: Aplicabilidade", 3.0)
        ade = row.get("Nota: Adequação", 3.0)
        
        ifrd = (0.25 * q) + (0.20 * ali) + (0.20 * apr) + (0.15 * rep) + (0.10 * app) + (0.10 * ade)
        return round(ifrd, 2)

    @staticmethod
    def get_classification(ifrd: float) -> str:
        """Determines the final classification text based on the IFRD score."""
        if ifrd >= 4.0:
            return "Bom Artigo"
        elif ifrd >= 3.0:
            return "Artigo Intermediário"
        else:
            return "Artigo Fraco"

    @classmethod
    def export(cls, articles_data: List[Dict[str, Any]], output_path: Path):
        """Processes raw articles data, computes indexes and exports to a stylized Excel sheet."""
        # Convert list of dicts to DataFrame
        df = pd.DataFrame(articles_data)

        # Calculate metrics for each row
        df["IFRD"] = df.apply(cls.calculate_ifrd, axis=1)
        df["Classificação Final"] = df["IFRD"].apply(cls.get_classification)

        # Reorder columns to place key info first
        column_order = [
            "Grupo", "Aluno", "Título", "Tem PDF", "IFRD", "Classificação Final",
            "Ano Publicação", "Editora", "Tipo Veículo", "Tipo Estudo",
            "Natureza Pesquisa", "Natureza Dados", "Tamanho Amostra",
            "Nota: Qualidade", "Nota: Replicabilidade", "Nota: Aplicabilidade",
            "Nota: Contribuição Teórica", "Nota: Adequação", "Nota: Aprendizagem", "Nota: Alinhamento",
            "Métodos Estatísticos", "Métricas Software",
            "Replicabilidade (Dados/Código)", "Link Replicação", "Elementos Compartilhados",
            "3 Principais Descobertas", "Principal Limitação", "Ameaças à Validade",
            "Unidades Plano", "Tópicos Específicos", "Conceito Ilustrado",
            "Resumo PT", "Justificativas"
        ]

        # Filter out order to match only columns that actually exist in df
        cols_to_use = [col for col in column_order if col in df.columns]
        # Append any remaining columns not listed in column_order
        cols_to_use.extend([col for col in df.columns if col not in cols_to_use])
        
        df_final = df[cols_to_use]

        # Write to Excel
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_final.to_excel(writer, sheet_name="Artigos Classificados", index=False)
            
            # Stylize the Excel worksheet
            workbook = writer.book
            worksheet = writer.sheets["Artigos Classificados"]

            # Styles setup
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Elegant Navy Blue
            
            # Fills for final classifications
            fill_good = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid") # Soft Green
            fill_mid = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Soft Yellow
            fill_bad = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")  # Soft Red
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )

            # Format headers
            for col_num, col_name in enumerate(df_final.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            # Format rows
            class_col_idx = df_final.columns.get_loc("Classificação Final") + 1
            ifrd_col_idx = df_final.columns.get_loc("IFRD") + 1
            
            for row_num in range(2, worksheet.max_row + 1):
                # Set default styles for all cells in the row
                for col_num in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = Font(name="Calibri", size=10)
                    cell.border = thin_border
                    
                    # Align number columns to center
                    if "Nota:" in str(df_final.columns[col_num-1]) or df_final.columns[col_num-1] in ["Tem PDF", "Ano Publicação", "IFRD"]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Color classification cells and IFRD cells based on the value
                class_cell = worksheet.cell(row=row_num, column=class_col_idx)
                ifrd_cell = worksheet.cell(row=row_num, column=ifrd_col_idx)
                
                val = str(class_cell.value)
                if "Bom" in val:
                    class_cell.fill = fill_good
                    ifrd_cell.fill = fill_good
                elif "Intermediário" in val:
                    class_cell.fill = fill_mid
                    ifrd_cell.fill = fill_mid
                elif "Fraco" in val:
                    class_cell.fill = fill_bad
                    ifrd_cell.fill = fill_bad

            # Auto-adjust column widths
            for col in worksheet.columns:
                col_letter = get_column_letter(col[0].column)
                # Cap the column width for long text columns
                col_name = str(worksheet.cell(row=1, column=col[0].column).value)
                if col_name in ["Resumo PT", "Justificativas", "3 Principais Descobertas", "Conceito Ilustrado", "Título"]:
                    worksheet.column_dimensions[col_letter].width = 40
                elif col_name in ["Métodos Estatísticos", "Métricas Software", "Ameaças à Validade", "Tópicos Específicos", "Principal Limitação"]:
                    worksheet.column_dimensions[col_letter].width = 25
                else:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

            # Freeze panes so header is always visible
            worksheet.freeze_panes = "A2"
