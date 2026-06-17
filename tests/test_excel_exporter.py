from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.discovery import ArticleMetadata
from src.excel_exporter import ExcelExporter
from src.models import ArticleResult, ProcessingStatus


def test_property_14_excel_export_contains_one_row_per_article_and_status(tmp_path):
    results = [
        ArticleResult(
            metadata=ArticleMetadata(
                folder_path=Path("artigos/a"),
                title="Artigo A",
                student="Ana",
                group="Turma",
                link="",
                pdf_path=Path("artigos/a/artigo.pdf"),
            ),
            processing_status=ProcessingStatus.SUCCESS,
            super_summary=None,
            classification=None,
            critic_output=None,
            citation_count=None,
            ifrd=None,
        ),
        ArticleResult(
            metadata=ArticleMetadata(
                folder_path=Path("artigos/b"),
                title="Artigo B",
                student="Bruno",
                group="Turma",
                link="",
                pdf_path=Path("artigos/b/artigo.pdf"),
            ),
            processing_status=ProcessingStatus.READER_FAILED,
            super_summary=None,
            classification=None,
            critic_output=None,
            citation_count=None,
            ifrd=None,
            error="falha",
        ),
    ]
    output_path = tmp_path / "out.xlsx"
    ExcelExporter.export(results, None, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook["Artigos Classificados"]
    headers = [cell.value for cell in sheet[1]]
    assert sheet.max_row == len(results) + 1
    assert "ProcessingStatus" in headers
    assert "Tema" in headers
    assert "Ano Publicação" in headers
    status_col = headers.index("ProcessingStatus") + 1
    assert [sheet.cell(row=row, column=status_col).value for row in range(2, sheet.max_row + 1)] == [
        "SUCCESS",
        "READER_FAILED",
    ]
    assert "Síntese Temática" in workbook.sheetnames
